"""
Import/Export utilities for Customer Master
Supports CSV, Excel, and PDF formats
"""
import csv
import logging
from io import BytesIO, StringIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth import get_user_model

from .models import CustomerMaster, KAMMaster
from .utils import generate_customer_number

logger = logging.getLogger(__name__)
User = get_user_model()


class CustomerExporter:
    """Export customer data to various formats"""
    
    @staticmethod
    def export_to_csv(queryset):
        """
        Export customers to CSV format
        Returns: HttpResponse with CSV file
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Header row
        headers = [
            'id',
            'customer_name',
            'nid',
            'company_name',
            'email',
            'phone',
            'address',
            'customer_type',
            'kam_id',
            'kam_designation',
            'customer_number',
            'contact_person',
            'status',
            'last_bill_invoice_date',
            'is_active',
            'created_at',
            'created_by',
            'updated_at',
            'updated_by',
        ]
        writer.writerow(headers)
        
        # Data rows
        for customer in queryset:
            writer.writerow([
                customer.id,
                customer.customer_name,
                customer.nid or '',
                customer.company_name or '',
                customer.email,
                customer.phone or '',
                customer.address,
                customer.customer_type,
                customer.kam_id.id if customer.kam_id else '',
                customer.kam_id.designation if customer.kam_id else '',
                customer.customer_number,
                customer.contact_person or '',
                customer.status,
                customer.last_bill_invoice_date or '',
                'Yes' if customer.is_active else 'No',
                customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                customer.created_by.id if customer.created_by else '',
                customer.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                customer.updated_by.id if customer.updated_by else '',
            ])
        
        return response
    
    @staticmethod
    def export_to_excel(queryset):
        """
        Export customers to Excel format with formatting
        Returns: HttpResponse with Excel file
        """
        # Create DataFrame
        data = []
        for customer in queryset:
            data.append({
                'id': customer.id,
                'customer_name': customer.customer_name,
                'nid': customer.nid or '',
                'company_name': customer.company_name or '',
                'email': customer.email,
                'phone': customer.phone or '',
                'address': customer.address,
                'customer_type': customer.customer_type,
                'kam_id': customer.kam_id.id if customer.kam_id else '',
                'kam_designation': customer.kam_id.designation if customer.kam_id else '',
                'customer_number': customer.customer_number,
                'contact_person': customer.contact_person or '',
                'status': customer.status,
                'last_bill_invoice_date': customer.last_bill_invoice_date or '',
                'is_active': 'Yes' if customer.is_active else 'No',
                'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'created_by': customer.created_by.id if customer.created_by else '',
                'updated_at': customer.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_by': customer.updated_by.id if customer.updated_by else '',
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel workbook with formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Customers', index=False)
            
            # Get worksheet
            worksheet = writer.sheets['Customers']
            
            # Format header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set column widths
            column_widths = {
                'A': 8,   # id
                'B': 20,  # customer_name
                'C': 18,  # nid
                'D': 18,  # company_name
                'E': 25,  # email
                'F': 15,  # phone
                'G': 20,  # address
                'H': 15,  # customer_type
                'I': 10,  # kam_id
                'J': 18,  # kam_designation
                'K': 18,  # customer_number
                'L': 15,  # contact_person
                'M': 12,  # status
                'N': 20,  # last_bill_invoice_date
                'O': 10,  # is_active
                'P': 20,  # created_at
                'Q': 12,  # created_by
                'R': 20,  # updated_at
                'S': 12,  # updated_by
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format data rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response


class CustomerImporter:
    """Import customer data from CSV or Excel formats"""

    @staticmethod
    def safe_int(value, default=0):
        """
        Safely convert value to int, treating empty strings as default
        """
        try:
            val = str(value).strip()
            return int(val) if val else default
        except (ValueError, TypeError):
            return default

    @staticmethod
    def safe_float(value, default=0.0):
        """
        Safely convert value to float, treating empty strings as default
        """
        try:
            val = str(value).strip()
            return float(val) if val else default
        except (ValueError, TypeError):
            return default

    @staticmethod
    def get_user_by_id(user_id):
        """
        Get user object by user ID
        Returns: User object or None
        """
        if not user_id or not str(user_id).strip():
            return None
        try:
            return User.objects.get(id=int(user_id))
        except (ValueError, User.DoesNotExist):
            return None
    
    @staticmethod
    def validate_required_fields(row_data, row_number):
        """
        Validate required fields in import data
        Returns: (is_valid, error_messages)
        """
        errors = []
        required_fields = ['customer_name', 'customer_type', 'kam_id']
        
        for field in required_fields:
            if not row_data.get(field, '').strip():
                errors.append(f"Row {row_number}: Missing required field '{field}'")
        
        # Validate email format
        email = row_data.get('email', '').strip()
        if email and '@' not in email:
            errors.append(f"Row {row_number}: Invalid email format '{email}'")
        
        # Validate customer type
        valid_types = dict(CustomerMaster.CUSTOMER_TYPE_CHOICES)
        customer_type = row_data.get('customer_type', '').strip().lower()
        if customer_type and customer_type not in valid_types:
            errors.append(f"Row {row_number}: Invalid customer_type '{customer_type}'. Must be one of: {', '.join(valid_types.keys())}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def import_from_csv(file_content):
        """
        Import customers from CSV file
        Returns: (success_count, error_messages, created_customers)
        """
        success_count = 0
        error_messages = []
        created_customers = []
        
        try:
            # Read CSV content
            csv_file = StringIO(file_content.decode('utf-8'))
            reader = csv.DictReader(csv_file)
            
            if not reader.fieldnames:
                return 0, ['Invalid CSV file: No headers found'], []
            
            for row_number, row in enumerate(reader, start=2):
                try:
                    # Validate row
                    is_valid, validation_errors = CustomerImporter.validate_required_fields(row, row_number)
                    if not is_valid:
                        error_messages.extend(validation_errors)
                        continue
                    
                    # Get KAM by ID
                    kam = None
                    kam_id = row.get('kam_id', '').strip()
                    if kam_id:
                        try:
                            kam = KAMMaster.objects.get(id=int(kam_id))
                        except (ValueError, KAMMaster.DoesNotExist):
                            error_messages.append(f"Row {row_number}: KAM with ID '{kam_id}' not found")
                            continue
                    
                    # Check if customer already exists (only if email is provided)
                    email = row.get('email', '').strip().lower()
                    if email and CustomerMaster.objects.filter(email=email).exists():
                        error_messages.append(f"Row {row_number}: Customer with email '{email}' already exists")
                        continue
                    
                    # Create customer
                    customer = CustomerMaster.objects.create(
                        customer_name=row.get('customer_name', '').strip(),
                        nid=row.get('nid', '').strip() or None,
                        company_name=row.get('company_name', '').strip() or None,
                        email=email,
                        phone=row.get('phone', '').strip() or '',
                        address=row.get('address', '').strip(),
                        customer_type=row.get('customer_type', '').strip().lower(),
                        kam_id=kam,
                        contact_person=row.get('contact_person', '').strip() or '',
                        status=row.get('status', 'active').lower(),
                        created_by=CustomerImporter.get_user_by_id(row.get('created_by', '')),
                        updated_by=CustomerImporter.get_user_by_id(row.get('updated_by', '')),
                    )
                    
                    # Auto-generate customer number
                    customer.customer_number = generate_customer_number(customer.customer_name, customer.id)
                    customer.save()
                    
                    # Prepare response with KAM and user details
                    customer_data = {
                        'id': customer.id,
                        'customer_name': customer.customer_name,
                        'nid': customer.nid,
                        'customer_number': customer.customer_number,
                        'email': customer.email,
                    }
                    if customer.kam_id:
                        customer_data['kam'] = {
                            'id': customer.kam_id.id,
                            'name': customer.kam_id.kam_name,
                            'designation': customer.kam_id.designation,
                            'email': customer.kam_id.email,
                        }
                    if customer.created_by:
                        customer_data['created_by'] = {
                            'id': customer.created_by.id,
                            'username': customer.created_by.username,
                            'email': customer.created_by.email,
                        }
                    if customer.updated_by:
                        customer_data['updated_by'] = {
                            'id': customer.updated_by.id,
                            'username': customer.updated_by.username,
                            'email': customer.updated_by.email,
                        }
                    created_customers.append(customer_data)
                    success_count += 1
                    
                except ValueError as e:
                    error_messages.append(f"Row {row_number}: Invalid data format - {str(e)}")
                except Exception as e:
                    error_messages.append(f"Row {row_number}: Error creating customer - {str(e)}")
        
        except Exception as e:
            error_messages.append(f"Error reading CSV file: {str(e)}")
        
        return success_count, error_messages, created_customers
    
    @staticmethod
    def import_from_excel(file_content):
        """
        Import customers from Excel file
        Returns: (success_count, error_messages, created_customers)
        """
        success_count = 0
        error_messages = []
        created_customers = []
        
        try:
            # Read Excel file
            df = pd.read_excel(BytesIO(file_content))
            
            if df.empty:
                return 0, ['Excel file is empty'], []
            
            for row_number, row in df.iterrows():
                try:
                    # Convert row to dictionary
                    row_data = row.to_dict()
                    
                    # Handle NaN values
                    row_data = {k: ('' if pd.isna(v) else str(v)) for k, v in row_data.items()}
                    
                    # Validate row
                    is_valid, validation_errors = CustomerImporter.validate_required_fields(row_data, row_number + 2)
                    if not is_valid:
                        error_messages.extend(validation_errors)
                        continue
                    
                    # Get KAM by ID
                    kam = None
                    kam_id = row_data.get('kam_id', '').strip()
                    if kam_id:
                        try:
                            kam = KAMMaster.objects.get(id=int(kam_id))
                        except (ValueError, KAMMaster.DoesNotExist):
                            error_messages.append(f"Row {row_number + 2}: KAM with ID '{kam_id}' not found")
                            continue
                    
                    # Check if customer already exists (only if email is provided)
                    email = row_data.get('email', '').strip().lower()
                    if email and CustomerMaster.objects.filter(email=email).exists():
                        error_messages.append(f"Row {row_number + 2}: Customer with email '{email}' already exists")
                        continue
                    
                    # Create customer
                    customer = CustomerMaster.objects.create(
                        customer_name=row_data.get('customer_name', '').strip(),
                        nid=row_data.get('nid', '').strip() or None,
                        company_name=row_data.get('company_name', '').strip() or None,
                        email=email,
                        phone=row_data.get('phone', '').strip() or '',
                        address=row_data.get('address', '').strip(),
                        customer_type=row_data.get('customer_type', '').strip().lower(),
                        kam_id=kam,
                        contact_person=row_data.get('contact_person', '').strip() or '',
                        status=row_data.get('status', 'active').lower(),
                        created_by=CustomerImporter.get_user_by_id(row_data.get('created_by', '')),
                        updated_by=CustomerImporter.get_user_by_id(row_data.get('updated_by', '')),
                    )
                    
                    # Auto-generate customer number
                    customer.customer_number = generate_customer_number(customer.customer_name, customer.id)
                    customer.save()
                    
                    # Prepare response with KAM and user details
                    customer_data = {
                        'id': customer.id,
                        'customer_name': customer.customer_name,
                        'nid': customer.nid,
                        'customer_number': customer.customer_number,
                        'email': customer.email,
                    }
                    if customer.kam_id:
                        customer_data['kam'] = {
                            'id': customer.kam_id.id,
                            'name': customer.kam_id.kam_name,
                            'designation': customer.kam_id.designation,
                            'email': customer.kam_id.email,
                        }
                    if customer.created_by:
                        customer_data['created_by'] = {
                            'id': customer.created_by.id,
                            'username': customer.created_by.username,
                            'email': customer.created_by.email,
                        }
                    if customer.updated_by:
                        customer_data['updated_by'] = {
                            'id': customer.updated_by.id,
                            'username': customer.updated_by.username,
                            'email': customer.updated_by.email,
                        }
                    created_customers.append(customer_data)
                    success_count += 1
                    
                except ValueError as e:
                    error_messages.append(f"Row {row_number + 2}: Invalid data format - {str(e)}")
                except Exception as e:
                    error_messages.append(f"Row {row_number + 2}: Error creating customer - {str(e)}")
        
        except Exception as e:
            error_messages.append(f"Error reading Excel file: {str(e)}")
        
        return success_count, error_messages, created_customers
