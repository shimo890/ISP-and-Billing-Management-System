"""
ISP-style KAM reporting: CSV, XLSX, PDF.
- Standardized KAM summary table (bill, collected, due, collection %).
- Per-KAM detail: service, customers, receivable ledger, trend, new/terminated.
- Bulk export: all KAMs in one workbook (multi-sheet) or one CSV with section blocks.
"""
import csv
import re
from io import BytesIO, StringIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFF9FAFB", size=10)
TITLE_FONT = Font(bold=True, size=14, color="FF111827")
SUB_FONT = Font(size=10, color="FF6B7280")
THIN = Side(style="thin", color="FFE5E7EB")


def _d(v):
    if v is None:
        return ""
    if isinstance(v, Decimal):
        return float(v)
    return v


def _pct(collected, bill):
    try:
        b = float(bill or 0)
        c = float(collected or 0)
        if b <= 0:
            return 0.0
        return round(100.0 * c / b, 1)
    except (TypeError, ValueError):
        return 0.0


def _due_by_kam_id(receivable_rows):
    """Aggregate closing due balance per KAM from ledger rows."""
    m = {}
    for r in receivable_rows or []:
        kid = r.get("kam_id")
        if kid is None:
            continue
        m[kid] = m.get(kid, 0.0) + float(r.get("total_due_balance") or 0)
    return m


def build_standard_kam_summary_rows(payload):
    """
    Enterprise ISP KAM summary: one row per KAM with bill, collected, due, rate.
    """
    overview = payload.get("kam_sales_overview") or []
    perf = payload.get("kam_performance") or []
    recv = payload.get("customer_receivable_rows") or []
    due_map = _due_by_kam_id(recv)
    perf_by_id = {p.get("kam_id"): p for p in perf if p.get("kam_id") is not None}
    rows = []
    for o in overview:
        kid = o.get("kam_id")
        p = perf_by_id.get(kid) or {}
        bill = float(o.get("total_sales_mrc") or 0)
        collected = float(p.get("total_revenue") or 0)
        due = due_map.get(kid)
        if due is None:
            due = max(0.0, bill - collected)
        rate = _pct(collected, bill)
        rows.append(
            {
                "kam_id": kid,
                "kam_name": o.get("kam_name") or "",
                "customers_count": o.get("customers_count") or 0,
                "total_capacity_mbps": float(o.get("total_capacity_mbps") or 0),
                "bill_mrc": bill,
                "collected": collected,
                "due_balance": float(due),
                "collection_rate_pct": rate,
                "termination_loss": float(p.get("termination_loss") or 0),
            }
        )
    rows.sort(key=lambda x: x["bill_mrc"], reverse=True)
    return rows


def _report_title_lines(filters_meta):
    f = filters_meta or {}
    return [
        "KTL / ISP — Sales & KAM Analytics Report",
        f"Reporting period: {f.get('start_date', '')} to {f.get('end_date', '')}",
    ]


def _append_detail_sections_csv(w, payload):
    """Append receivable, trend, new/term for KAM detail exports."""
    kr = payload.get("kam_receivable_rows") or []
    if kr:
        w.writerow([])
        w.writerow(["Customer receivable (ledger)"])
        w.writerow(
            [
                "Customer",
                "Opening",
                "Bill",
                "Received",
                "Due",
                "Status",
            ]
        )
        for r in kr:
            w.writerow(
                [
                    r.get("customer_name"),
                    _d(r.get("opening_balance")),
                    _d(r.get("total_bill_amount")),
                    _d(r.get("total_payment_received")),
                    _d(r.get("total_due_balance")),
                    r.get("status_display") or r.get("status"),
                ]
            )

    trend = payload.get("kam_monthly_trend") or []
    if trend:
        w.writerow([])
        w.writerow(["Monthly trend (bill vs collected)"])
        w.writerow(["Month", "Bill", "Collected"])
        for t in trend:
            w.writerow([t.get("label"), _d(t.get("bill")), _d(t.get("collected"))])

    new_k = payload.get("kam_new_customers") or []
    term_k = payload.get("kam_terminated_customers") or []
    if new_k or term_k:
        w.writerow([])
        w.writerow(["New customers (period)"])
        w.writerow(["Customer", "MRC"])
        for c in new_k:
            w.writerow([c.get("customer_name"), _d(c.get("mrc"))])
        w.writerow([])
        w.writerow(["Terminated / discontinued"])
        w.writerow(["Customer", "Revenue loss", "Date"])
        for c in term_k:
            w.writerow(
                [
                    c.get("customer_name"),
                    _d(c.get("revenue_loss")),
                    str(c.get("termination_date") or ""),
                ]
            )


def build_sales_export_csv(
    payload,
    kam_drill=None,
    include_executive_summary=True,
    include_banner=True,
):
    """UTF-8 CSV with BOM for Excel."""
    buf = StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf)

    if include_banner:
        for line in _report_title_lines(payload.get("filters")):
            w.writerow([line])
        w.writerow([])

    rs = payload.get("revenue_summary") or {}
    if include_executive_summary:
        w.writerow(["Executive summary — revenue"])
        w.writerow(["Total active MRC (portfolio)", _d(rs.get("total_active_mrc"))])
        w.writerow(["New MRC this period", _d(rs.get("new_mrc_this_period"))])
        w.writerow(["Lost MRC (terminations)", _d(rs.get("lost_mrc"))])
        w.writerow(["Net growth (new − lost)", _d(rs.get("net_growth"))])
        w.writerow([])

    std_rows = build_standard_kam_summary_rows(payload)
    if std_rows:
        w.writerow(["Standardized KAM performance summary"])
        w.writerow(
            [
                "KAM",
                "KAM ID",
                "Customers",
                "Capacity (Mbps)",
                "Bill MRC",
                "Collected",
                "Due balance",
                "Collection %",
                "Term. loss",
            ]
        )
        for r in std_rows:
            w.writerow(
                [
                    r["kam_name"],
                    r["kam_id"],
                    r["customers_count"],
                    round(r["total_capacity_mbps"], 2),
                    round(r["bill_mrc"], 2),
                    round(r["collected"], 2),
                    round(r["due_balance"], 2),
                    r["collection_rate_pct"],
                    round(r["termination_loss"], 2),
                ]
            )
        w.writerow([])

    svc_over = payload.get("service_sales_overview") or []
    if svc_over:
        w.writerow(["Sales by service (portfolio)"])
        w.writerow(["Service", "Mbps", "Qty (lines)", "MRC"])
        for r in svc_over:
            w.writerow(
                [
                    r.get("service_type"),
                    _d(r.get("total_mbps")),
                    r.get("qty"),
                    _d(r.get("total_mrc")),
                ]
            )
        w.writerow([])

    kd = payload.get("kam_detail")
    if kd:
        w.writerow(["KAM detail — service breakdown", kd.get("kam_name")])
        w.writerow(["Service", "Mbps", "Qty (lines)", "MRC"])
        for row in kd.get("by_service") or []:
            w.writerow(
                [
                    row.get("service_type"),
                    _d(row.get("total_mbps")),
                    row.get("qty", row.get("service_lines")),
                    _d(row.get("total_mrc")),
                ]
            )
        w.writerow(["Total capacity (Mbps)", _d(kd.get("total_capacity_mbps"))])
        w.writerow(["Total MRC", _d(kd.get("total_mrc"))])
        w.writerow([])

    kc = payload.get("kam_customers") or []
    if kc:
        w.writerow(["Customers under KAM (account summary)"])
        w.writerow(["Customer", "Company", "Services", "Mbps", "Qty (lines)", "MRC", "Status"])
        for c in kc:
            w.writerow(
                [
                    c.get("customer_name"),
                    c.get("company_name"),
                    c.get("service_type"),
                    _d(c.get("capacity_mbps")),
                    c.get("qty"),
                    _d(c.get("mrc")),
                    c.get("status_display"),
                ]
            )
        w.writerow([])

    if kam_drill:
        _append_detail_sections_csv(w, payload)

    return buf.getvalue().encode("utf-8")


def _style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_standard_kam_table_xlsx(ws, start_row, payload):
    """Write standardized KAM table; returns next row index."""
    r = start_row
    ws.cell(r, 1, "Standardized KAM performance summary").font = TITLE_FONT
    r += 1
    headers = [
        "KAM",
        "KAM ID",
        "Customers",
        "Capacity (Mbps)",
        "Bill MRC",
        "Collected",
        "Due balance",
        "Collection %",
        "Term. loss",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h)
    _style_header_row(ws, r, len(headers))
    r += 1
    std_rows = build_standard_kam_summary_rows(payload)
    for row in std_rows:
        ws.cell(r, 1, row["kam_name"])
        ws.cell(r, 2, row["kam_id"])
        ws.cell(r, 3, row["customers_count"])
        ws.cell(r, 4, float(row["total_capacity_mbps"]))
        ws.cell(r, 5, float(row["bill_mrc"]))
        ws.cell(r, 6, float(row["collected"]))
        ws.cell(r, 7, float(row["due_balance"]))
        ws.cell(r, 8, float(row["collection_rate_pct"]))
        ws.cell(r, 9, float(row["termination_loss"]))
        r += 1
    return r + 1


def _write_kam_detail_blocks_xlsx(ws, start_row, payload):
    """Service, customers, receivable, trend, new/term on one sheet."""
    r = start_row
    kd = payload.get("kam_detail") or {}
    if kd.get("kam_name"):
        ws.cell(r, 1, f"KAM: {kd.get('kam_name')}").font = TITLE_FONT
        r += 2

    if kd.get("by_service"):
        ws.cell(r, 1, "Service breakdown").font = Font(bold=True, size=11)
        r += 1
        for c, h in enumerate(["Service", "Mbps", "Qty (lines)", "MRC"], 1):
            ws.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws, r, 4)
        r += 1
        for row in kd.get("by_service") or []:
            ws.cell(r, 1, row.get("service_type"))
            ws.cell(r, 2, float(row.get("total_mbps") or 0))
            ws.cell(r, 3, row.get("qty", row.get("service_lines")))
            ws.cell(r, 4, float(row.get("total_mrc") or 0))
            r += 1
        r += 1

    kc = payload.get("kam_customers") or []
    if kc:
        ws.cell(r, 1, "Account list (capacity & MRC)").font = Font(bold=True, size=11)
        r += 1
        hdr = ["Customer", "Company", "Services", "Mbps", "Qty", "MRC", "Status"]
        for c, h in enumerate(hdr, 1):
            ws.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws, r, len(hdr))
        r += 1
        for c in kc:
            ws.cell(r, 1, c.get("customer_name"))
            ws.cell(r, 2, c.get("company_name"))
            ws.cell(r, 3, c.get("service_type"))
            ws.cell(r, 4, float(c.get("capacity_mbps") or 0))
            ws.cell(r, 5, c.get("qty") or 0)
            ws.cell(r, 6, float(c.get("mrc") or 0))
            ws.cell(r, 7, c.get("status_display"))
            r += 1
        r += 1

    kr = payload.get("kam_receivable_rows") or []
    if kr:
        ws.cell(r, 1, "Receivable ledger (period)").font = Font(bold=True, size=11)
        r += 1
        hdr = ["Customer", "Opening", "Bill", "Received", "Due", "Coll. %", "Status"]
        for c, h in enumerate(hdr, 1):
            ws.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws, r, len(hdr))
        r += 1
        for row in kr:
            bill = float(row.get("total_bill_amount") or 0)
            rcvd = float(row.get("total_payment_received") or 0)
            cr = int(round(100 * rcvd / bill)) if bill > 0 else 0
            ws.cell(r, 1, row.get("customer_name"))
            ws.cell(r, 2, float(row.get("opening_balance") or 0))
            ws.cell(r, 3, bill)
            ws.cell(r, 4, rcvd)
            ws.cell(r, 5, float(row.get("total_due_balance") or 0))
            ws.cell(r, 6, cr)
            ws.cell(r, 7, row.get("status_display") or row.get("status"))
            r += 1
        r += 1

    trend = payload.get("kam_monthly_trend") or []
    if trend:
        ws.cell(r, 1, "Monthly trend").font = Font(bold=True, size=11)
        r += 1
        for c, h in enumerate(["Month", "Bill", "Collected"], 1):
            ws.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws, r, 3)
        r += 1
        for t in trend:
            ws.cell(r, 1, t.get("label"))
            ws.cell(r, 2, float(t.get("bill") or 0))
            ws.cell(r, 3, float(t.get("collected") or 0))
            r += 1
        r += 1

    new_k = payload.get("kam_new_customers") or []
    term_k = payload.get("kam_terminated_customers") or []
    if new_k:
        ws.cell(r, 1, "New customers").font = Font(bold=True, size=11)
        r += 1
        ws.cell(r, 1, "Customer").font = Font(bold=True)
        ws.cell(r, 2, "MRC").font = Font(bold=True)
        r += 1
        for c in new_k:
            ws.cell(r, 1, c.get("customer_name"))
            ws.cell(r, 2, float(c.get("mrc") or 0))
            r += 1
        r += 1
    if term_k:
        ws.cell(r, 1, "Terminated").font = Font(bold=True, size=11)
        r += 1
        for c, h in enumerate(["Customer", "Loss", "Date"], 1):
            ws.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws, r, 3)
        r += 1
        for c in term_k:
            ws.cell(r, 1, c.get("customer_name"))
            ws.cell(r, 2, float(c.get("revenue_loss") or 0))
            ws.cell(r, 3, str(c.get("termination_date") or ""))
            r += 1

    return r


def build_sales_export_xlsx(payload, kam_drill=None):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"
    row = 1
    for line in _report_title_lines(payload.get("filters")):
        ws0.cell(row, 1, line).font = TITLE_FONT if row == 1 else SUB_FONT
        row += 1
    row += 1

    rs = payload.get("revenue_summary") or {}
    ws0.cell(row, 1, "Executive summary").font = Font(bold=True, size=12)
    row += 1
    for label, key in [
        ("Total active MRC (portfolio)", "total_active_mrc"),
        ("New MRC this period", "new_mrc_this_period"),
        ("Lost MRC (terminations)", "lost_mrc"),
        ("Net growth", "net_growth"),
    ]:
        ws0.cell(row, 1, label)
        ws0.cell(row, 2, float(rs.get(key) or 0))
        row += 1
    row += 1

    row = _write_standard_kam_table_xlsx(ws0, row, payload)

    svc_over = payload.get("service_sales_overview") or []
    if svc_over:
        ws0.cell(row, 1, "Sales by service (portfolio)").font = Font(bold=True, size=12)
        row += 1
        for c, h in enumerate(["Service", "Mbps", "Qty (lines)", "MRC"], 1):
            ws0.cell(row, c, h).font = Font(bold=True)
        _style_header_row(ws0, row, 4)
        row += 1
        for r in svc_over:
            ws0.cell(row, 1, r.get("service_type"))
            ws0.cell(row, 2, float(r.get("total_mbps") or 0))
            ws0.cell(row, 3, r.get("qty") or 0)
            ws0.cell(row, 4, float(r.get("total_mrc") or 0))
            row += 1
        row += 1

    kd = payload.get("kam_detail")
    if kd:
        ws1 = wb.create_sheet("KAM detail")
        ws1.cell(1, 1, f"KAM: {kd.get('kam_name')}").font = TITLE_FONT
        r = 3
        for c, h in enumerate(["Service", "Mbps", "Qty (lines)", "MRC"], 1):
            ws1.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws1, r, 4)
        r += 1
        for svc in kd.get("by_service") or []:
            ws1.cell(r, 1, svc.get("service_type"))
            ws1.cell(r, 2, float(svc.get("total_mbps") or 0))
            ws1.cell(r, 3, svc.get("qty", svc.get("service_lines")))
            ws1.cell(r, 4, float(svc.get("total_mrc") or 0))
            r += 1
        r += 1
        ws1.cell(r, 1, "Total capacity (Mbps)").font = Font(bold=True)
        ws1.cell(r, 2, float(kd.get("total_capacity_mbps") or 0))
        r += 1
        ws1.cell(r, 1, "Total MRC").font = Font(bold=True)
        ws1.cell(r, 2, float(kd.get("total_mrc") or 0))

    kc = payload.get("kam_customers") or []
    if kc:
        ws2 = wb.create_sheet("Customers")
        headers = ["Customer", "Company", "Services", "Mbps", "Qty", "MRC", "Status"]
        for c, h in enumerate(headers, 1):
            ws2.cell(1, c, h).font = Font(bold=True)
        _style_header_row(ws2, 1, len(headers))
        for i, cust in enumerate(kc, 2):
            ws2.cell(i, 1, cust.get("customer_name"))
            ws2.cell(i, 2, cust.get("company_name"))
            ws2.cell(i, 3, cust.get("service_type"))
            ws2.cell(i, 4, float(cust.get("capacity_mbps") or 0))
            ws2.cell(i, 5, cust.get("qty") or 0)
            ws2.cell(i, 6, float(cust.get("mrc") or 0))
            ws2.cell(i, 7, cust.get("status_display"))

    if kam_drill and (
        payload.get("kam_receivable_rows")
        or payload.get("kam_monthly_trend")
        or payload.get("kam_new_customers")
        or payload.get("kam_terminated_customers")
    ):
        ws3 = wb.create_sheet("Receivable & trend")
        _write_kam_detail_blocks_xlsx(ws3, 1, payload)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_sales_export_pdf(payload, kam_drill=None):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("ISP KAM — Sales Analytics Report", styles["Title"]))
    f = payload.get("filters") or {}
    story.append(
        Paragraph(
            f"Period: {f.get('start_date', '')} to {f.get('end_date', '')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))

    rs = payload.get("revenue_summary") or {}
    story.append(Paragraph("Executive summary", styles["Heading2"]))
    rev_data = [
        ["Metric", "Amount"],
        ["Total Active MRC", str(_d(rs.get("total_active_mrc")))],
        ["New MRC", str(_d(rs.get("new_mrc_this_period")))],
        ["Lost MRC", str(_d(rs.get("lost_mrc")))],
        ["Net Growth", str(_d(rs.get("net_growth")))],
    ]
    t = Table(rev_data, colWidths=[3 * inch, 2 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 16))

    std_rows = build_standard_kam_summary_rows(payload)
    if std_rows:
        story.append(Paragraph("Standardized KAM summary", styles["Heading2"]))
        data = [["KAM", "Cust", "Mbps", "Bill", "Paid", "Due", "%"]]
        for r in std_rows[:40]:
            data.append(
                [
                    str(r["kam_name"])[:28],
                    str(r["customers_count"]),
                    str(round(r["total_capacity_mbps"], 0)),
                    str(round(r["bill_mrc"], 0)),
                    str(round(r["collected"], 0)),
                    str(round(r["due_balance"], 0)),
                    str(r["collection_rate_pct"]),
                ]
            )
        ts = Table(data, repeatRows=1, colWidths=[2 * inch, 0.45 * inch, 0.65 * inch, 0.85 * inch, 0.85 * inch, 0.75 * inch, 0.4 * inch])
        ts.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(ts)
        story.append(Spacer(1, 12))

    svc_over = payload.get("service_sales_overview") or []
    if svc_over:
        story.append(Paragraph("Sales by service", styles["Heading2"]))
        data = [["Service", "Mbps", "Qty", "MRC"]]
        for r in svc_over[:25]:
            data.append(
                [
                    str(r.get("service_type") or "")[:35],
                    str(_d(r.get("total_mbps"))),
                    str(r.get("qty") or ""),
                    str(_d(r.get("total_mrc"))),
                ]
            )
        ts2 = Table(data, repeatRows=1)
        ts2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(ts2)
        story.append(Spacer(1, 16))

    kd = payload.get("kam_detail")
    if kd:
        story.append(PageBreak())
        story.append(Paragraph(f"KAM detail — {kd.get('kam_name')}", styles["Heading2"]))
        data = [["Service", "Mbps", "Qty", "MRC"]]
        for row in kd.get("by_service") or []:
            data.append(
                [
                    str(row.get("service_type") or ""),
                    str(_d(row.get("total_mbps"))),
                    str(row.get("qty", row.get("service_lines"))),
                    str(_d(row.get("total_mrc"))),
                ]
            )
        t3 = Table(data, repeatRows=1)
        t3.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(t3)

    doc.build(story)
    return buf.getvalue()


def _sanitize_sheet_name(name, kid, used):
    """Excel sheet names: max 31 chars, no []:*?/\\"""
    raw = re.sub(r"[\[\]:*?/\\]", " ", str(name or "").strip())
    base = (raw[:28] if raw else f"KAM_{kid}").strip() or f"KAM_{kid}"
    candidate = base[:31]
    n = 0
    while candidate in used:
        n += 1
        candidate = f"{base[:20]}_{kid}_{n}"[:31]
    used.add(candidate)
    return candidate


def dec_to_float(obj):
    """Match SalesAnalyticsExportView serialization."""
    if hasattr(obj, "quantize"):
        return float(obj)
    if isinstance(obj, dict):
        return {k: dec_to_float(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [dec_to_float(x) for x in obj]
    return obj


def build_bulk_kam_export_xlsx(service):
    """
    One workbook: Summary sheet + one sheet per KAM with full detail blocks.
    `service` is SalesAnalyticsService instance (date-scoped).
    """
    overview = dec_to_float(service.get_full_analytics(mode="overview"))
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "KAM Summary"
    row = 1
    for line in _report_title_lines(overview.get("filters")):
        ws0.cell(row, 1, line).font = TITLE_FONT if row == 1 else SUB_FONT
        row += 1
    row += 1
    rs = overview.get("revenue_summary") or {}
    ws0.cell(row, 1, "Executive summary").font = Font(bold=True, size=12)
    row += 1
    for label, key in [
        ("Total active MRC (portfolio)", "total_active_mrc"),
        ("New MRC this period", "new_mrc_this_period"),
        ("Lost MRC (terminations)", "lost_mrc"),
        ("Net growth", "net_growth"),
    ]:
        ws0.cell(row, 1, label)
        ws0.cell(row, 2, float(rs.get(key) or 0))
        row += 1
    row += 1
    _write_standard_kam_table_xlsx(ws0, row, overview)

    svc_over = overview.get("service_sales_overview") or []
    if svc_over:
        ws_svc = wb.create_sheet("Service mix")
        ws_svc.cell(1, 1, "Sales by service (all KAMs)").font = TITLE_FONT
        r = 3
        for c, h in enumerate(["Service", "Mbps", "Qty (lines)", "MRC"], 1):
            ws_svc.cell(r, c, h).font = Font(bold=True)
        _style_header_row(ws_svc, r, 4)
        r += 1
        for sv in svc_over:
            ws_svc.cell(r, 1, sv.get("service_type"))
            ws_svc.cell(r, 2, float(sv.get("total_mbps") or 0))
            ws_svc.cell(r, 3, sv.get("qty") or 0)
            ws_svc.cell(r, 4, float(sv.get("total_mrc") or 0))
            r += 1

    used_names = {"KAM Summary", "Service mix"}
    for o in overview.get("kam_sales_overview") or []:
        kid = o.get("kam_id")
        if not kid:
            continue
        detail = dec_to_float(service.get_full_analytics(kam_id_drill=kid, mode="detail"))
        name = (detail.get("kam_detail") or {}).get("kam_name") or o.get("kam_name") or str(kid)
        sheet_name = _sanitize_sheet_name(name, kid, used_names)
        ws = wb.create_sheet(sheet_name)
        rr = 1
        for line in _report_title_lines(detail.get("filters")):
            ws.cell(rr, 1, f"{line} — {name}").font = TITLE_FONT if rr == 1 else SUB_FONT
            rr += 1
        pr = detail.get("kam_performance_row") or {}
        ws.cell(rr, 1, "Period totals").font = Font(bold=True, size=11)
        rr += 1
        ws.cell(rr, 1, "Bill MRC")
        ws.cell(rr, 2, float(pr.get("total_sales_mrc") or 0))
        rr += 1
        ws.cell(rr, 1, "Collected")
        ws.cell(rr, 2, float(pr.get("total_revenue") or 0))
        rr += 1
        ws.cell(rr, 1, "Due (ledger)")
        ws.cell(rr, 2, float(detail.get("kam_total_due_ledger") or 0))
        rr += 1
        ws.cell(rr, 1, "Opening (est.)")
        ws.cell(rr, 2, float(detail.get("kam_total_opening") or 0))
        rr += 2
        _write_kam_detail_blocks_xlsx(ws, rr, detail)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_bulk_kam_export_csv(service):
    """Single CSV: portfolio summary, then one block per KAM (detail sections)."""
    overview = dec_to_float(service.get_full_analytics(mode="overview"))
    buf = StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf)

    w.writerow(["BULK EXPORT — ALL KAMS (ISP KAM REPORT)"])
    for line in _report_title_lines(overview.get("filters")):
        w.writerow([line])
    w.writerow([])

    w.writerow(["=== PORTFOLIO SUMMARY ==="])
    w.writerow([])
    main_part = build_sales_export_csv(overview, kam_drill=None, include_banner=False)
    text = main_part.decode("utf-8-sig")
    for line in text.splitlines():
        if not line.strip():
            w.writerow([])
        else:
            w.writerow(next(csv.reader([line])))

    for o in overview.get("kam_sales_overview") or []:
        kid = o.get("kam_id")
        if not kid:
            continue
        detail = dec_to_float(service.get_full_analytics(kam_id_drill=kid, mode="detail"))
        name = (detail.get("kam_detail") or {}).get("kam_name") or o.get("kam_name") or str(kid)
        w.writerow([])
        w.writerow([f"=== KAM DETAIL: {name} (ID {kid}) ==="])
        w.writerow([])
        detail_part = build_sales_export_csv(
            detail,
            kam_drill=kid,
            include_executive_summary=False,
            include_banner=False,
        )
        dtext = detail_part.decode("utf-8-sig")
        for line in dtext.splitlines():
            if not line.strip():
                w.writerow([])
            else:
                w.writerow(next(csv.reader([line])))

    return buf.getvalue().encode("utf-8")
